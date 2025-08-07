import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os

# Define the function to process each file
def process_file(file_path):
    print(f"Start: '{os.path.basename(file_path)}'", flush=True)
    
    try:
        # Load the original data starting from row 2 (pandas uses zero-indexing, so header=1)
        print("Loading data...", flush=True)
        df = pd.read_excel(file_path, header=0)  # Use the 2nd row as header

        # Load the top 1 row separately (assumed to be title or header information)
        print("Loading top row...", flush=True)
        top_rows = pd.read_excel(file_path, nrows=1, header=None)

        # Convert the timestamp column to datetime format
        print("Converting timestamps...", flush=True)
        df['timestamp'] = pd.to_datetime(df['timestamp'])

        # Create a complete timestamp range with the same frequency as in the data
        print("Creating complete timestamp range...", flush=True)
        start_time = df['timestamp'].min()
        end_time = df['timestamp'].max()
        all_timestamps = pd.date_range(start=start_time, end=end_time, freq='T')  # Assuming frequency is 1 minute

        # Find missing timestamps
        print("Finding missing timestamps...", flush=True)
        missing_timestamps = all_timestamps.difference(df['timestamp'])

        # Create a DataFrame for missing timestamps
        print("Creating DataFrame for missing timestamps...", flush=True)
        missing_df = pd.DataFrame(missing_timestamps, columns=['timestamp'])
        missing_df['energy rate'] = 0  # Add the same column as in the original data

        # Combine the original and missing data
        print("Combining original and missing data...", flush=True)
        combined_df = pd.concat([df, missing_df]).sort_values(by='timestamp').reset_index(drop=True)

        # Ensure the output directory exists
        print("Ensuring output directory exists...", flush=True)
        input_dir, input_file = os.path.split(file_path)
        parent_dir = os.path.dirname(input_dir)
        output_dir = os.path.join(parent_dir, "time corrected")
        os.makedirs(output_dir, exist_ok=True)

        # Generate the output file path
        print("Generating output file path...", flush=True)
        output_file_name = f"time corrected {input_file}"
        output_file_path = os.path.join(output_dir, output_file_name)

        # Save the combined data back to an Excel file
        print("Saving combined data to Excel...", flush=True)
        with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
            top_rows.to_excel(writer, index=False, header=False)
            # Write combined data starting from row 2 (row index 1 in Excel) to preserve the top row
            combined_df.to_excel(writer, startrow=0, index=False)
        print(f"Saved: '{os.path.basename(output_file_path)}'", flush=True)

        # Load the workbook and select the active worksheet
        print("Loading workbook for highlighting...", flush=True)
        wb = load_workbook(output_file_path)
        ws = wb.active

        # Define a red fill
        red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

        # Get the max column once
        max_column = ws.max_column

        # Highlight the new rows in red
        print("Highlighting new rows in red...", flush=True)
        # Since the top row occupies the first row, the data starts on row 2, so add an offset of 1.
        for index, row in combined_df.iterrows():
            if row['timestamp'] in missing_timestamps:
                row_idx = index + 2  # Adjusting for the top row
                for col_idx in range(1, max_column + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = red_fill

        # Save the workbook with highlighted rows
        print("Saving highlighted workbook...", flush=True)
        wb.save(output_file_path)

        # Log the missing timestamps
        print("Logging missing timestamps...", flush=True)
        log_file_path = os.path.join(output_dir, "timestamp_log.csv")
        log_exists = os.path.isfile(log_file_path)

        with open(log_file_path, 'a') as log_file:
            if not log_exists:
                log_file.write("File Name, Missing Timestamps Count, Missing Timestamps\n")
            log_file.write(f"{input_file}, {len(missing_timestamps)}, {'; '.join(missing_timestamps.astype(str))}\n")

        print(f"Log updated: '{os.path.basename(log_file_path)}'", flush=True)
    
    except Exception as e:
        print(f"Error processing '{os.path.basename(file_path)}': {e}", flush=True)

# Automatically find all .xlsx files in the "original" folder
print("Finding all .xlsx files in the 'original' folder...", flush=True)
current_dir = os.path.dirname(os.path.abspath(__file__))
original_folder = os.path.join(os.path.dirname(current_dir), "original")
file_paths = [os.path.join(original_folder, f) for f in os.listdir(original_folder) if f.endswith('.xlsx')]

# Process files one at a time
for file_path in file_paths:
    process_file(file_path)

print("All files processed.", flush=True)
