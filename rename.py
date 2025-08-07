import openpyxl
import os
from concurrent.futures import ThreadPoolExecutor

def sanitize_filename(filename):
    # Replace invalid characters with underscores
    return filename.replace(':', '_').replace(' ', '_').replace('\\', '_').replace('/', '_')

def process_file(file_path):
    try:
        # Load the workbook and select the active sheet
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        # Extract values from cells B2 and B5
        value_b2 = ws['B2'].value
        value_b5 = ws['B5'].value

        # Extract first and last values in the first column, excluding the first 5 rows
        first_value = None
        last_value = None
        for row in ws.iter_rows(min_row=6, min_col=1, max_col=1):
            cell_value = row[0].value
            if cell_value is not None:
                if first_value is None:
                    first_value = cell_value.date() if hasattr(cell_value, 'date') else cell_value
                last_value = cell_value.date() if hasattr(cell_value, 'date') else cell_value

        # Full debug print statements
        print(f"Processing File: {os.path.basename(file_path)}", flush=True)
        print(f"Extracted Values:", flush=True)
        print(f"  B2: {value_b2}", flush=True)
        print(f"  B5: {value_b5}", flush=True)
        print(f"  First Value in Column A (after row 5): {first_value}", flush=True)
        print(f"  Last Value in Column A (after row 5): {last_value}", flush=True)

        # Construct the new file name
        folder_path = os.path.dirname(file_path)
        new_filename = sanitize_filename(f"{value_b2} {value_b5} {first_value} {last_value}.xlsx")
        new_file_path = os.path.join(folder_path, new_filename)

        # Rename the original file
        os.rename(file_path, new_file_path)
        print(f"Renamed to: {new_filename}", flush=True)

        # Now reopen the renamed workbook to delete the top 4 rows
        wb2 = openpyxl.load_workbook(new_file_path)
        ws2 = wb2.active

        # Delete rows 1 through 4
        ws2.delete_rows(1, 4)

        # Save the workbook (overwrite)
        wb2.save(new_file_path)
        print(f"Deleted top 4 rows and saved: {new_filename}\n", flush=True)

    except Exception as e:
        print(f"Error processing {os.path.basename(file_path)}: {e}", flush=True)

def process_files(file_paths):
    with ThreadPoolExecutor() as executor:
        executor.map(process_file, file_paths)

if __name__ == "__main__":
    # Define the path to the "original" folder
    script_dir = os.path.dirname(os.path.abspath(__file__))
    original_folder_path = os.path.join(script_dir, '..', 'original')

    # Get all .xlsx files in the "original" folder
    file_paths = [
        os.path.join(original_folder_path, f)
        for f in os.listdir(original_folder_path)
        if f.lower().endswith('.xlsx')
    ]

    if file_paths:
        print(f"Found {len(file_paths)} .xlsx files to process.\n", flush=True)
        process_files(file_paths)
    else:
        print("No .xlsx files found", flush=True)
