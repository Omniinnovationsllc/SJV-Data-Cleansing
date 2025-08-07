import openpyxl
import os
from concurrent.futures import ThreadPoolExecutor, as_completed

def process_excel(file_path):
    print(f"Start: '{os.path.basename(file_path)}'", flush=True)
    
    # Load the workbook and select the active sheet
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    # Get the maximum row number
    max_row = sheet.max_row

    # Store the indices of cells to be processed
    cells_to_process = []

    for row in range(7, max_row + 1):
        cell = sheet.cell(row=row, column=2)
        fill = cell.fill

        if fill.start_color.index == "FFFF0000":  # This is the red fill color in RGB
            cells_to_process.append(row)

    for row in cells_to_process:
        # Find the first non-red value before the cell
        prev_value = None
        for r in range(row - 1, 6, -1):
            cell = sheet.cell(row=r, column=2)
            fill = cell.fill
            if fill.start_color.index != "FFFF0000":
                prev_value = cell.value
                break

        # Find the first non-red value after the cell
        next_value = None
        for r in range(row + 1, max_row + 1):
            cell = sheet.cell(row=r, column=2)
            fill = cell.fill
            if fill.start_color.index != "FFFF0000":
                next_value = cell.value
                break

        # Calculate the average
        if prev_value is not None and next_value is not None:
            new_value = (prev_value + next_value) / 2
            sheet.cell(row=row, column=2).value = new_value

    # Create the 'averaged' directory one level up from the script's location if it doesn't exist
    script_dir = os.path.dirname(os.path.abspath(__file__))
    parent_dir = os.path.dirname(script_dir)
    averaged_dir = os.path.join(parent_dir, "averaged")
    if not os.path.exists(averaged_dir):
        os.makedirs(averaged_dir)

    # Modify the output file name
    input_file_name = os.path.basename(file_path)
    output_file_name = input_file_name.replace("time corrected", "averaged")
    output_file_path = os.path.join(averaged_dir, output_file_name)

    # Save the workbook
    wb.save(output_file_path)
    print(f"Saved: '{output_file_path}'", flush=True)

def main():
    print("Conversion start...", flush=True)
    
    # Set the directory to the 'time corrected' folder one level up from the script's directory
    script_dir = os.path.dirname(os.path.abspath(__file__))
    parent_dir = os.path.dirname(script_dir)
    time_corrected_dir = os.path.join(parent_dir, "time corrected")

    # Get all .xlsx files in the 'time corrected' directory
    file_paths = [os.path.join(time_corrected_dir, f) for f in os.listdir(time_corrected_dir) if f.endswith(".xlsx")]

    if file_paths:
        print(f"Files found: {len(file_paths)}", flush=True)
        # Use ThreadPoolExecutor to process files concurrently
        with ThreadPoolExecutor() as executor:
            futures = {executor.submit(process_excel, file_path): file_path for file_path in file_paths}

            for future in as_completed(futures):
                file_path = futures[future]
                try:
                    future.result()
                except Exception as e:
                    print(f"Error: '{file_path}': {e}", flush=True)
    else:
        print("No files found", flush=True)

if __name__ == "__main__":
    main()
    print("Conversion done.", flush=True)
